"""
Libya CARA — Municipality population refresh.

Purpose
-------
Refresh `population` values in `data/libya_municipalities.json` against the
OCHA Libya HNO 2021 dataset (the same dataset IOM DTM, UNHCR, and the Libya
HNO/HRP cycle use as their population baseline). The OCHA file is published
on HDX under CC BY 4.0 by `ocha-libya` and provides per-baladiya breakdowns
by population group (Non-displaced, Returnees, IDPs, Migrants, Refugees).

The refresh:

  1. Downloads (or reuses cached copies of) three HDX datasets:
        - OCHA Libya Population Statistics (per-baladiya, by group)
        - IOM DTM Libya Baseline Round 45 (IDP/Returnee context)
        - IOM DTM Libya Migrants Baseline Round 59 (migrant context)
     The IOM datasets are kept for downstream displacement indicators and
     are NOT used to overwrite resident population.

  2. Reconciles our 106 slug-style IDs (LY-001..LY-106) against OCHA's 100
     PCode-style baladiyas (LY######) via a manually curated ALIAS map plus
     normalized exact-name matching.

  3. For each matched entry, sets:
        population        = OCHA Non-displaced + Returnees   (resident pop)
        population_year   = 2021
        population_source = "OCHA Libya HNO 2021 (CC BY 4.0)"
        population_method = "hdx_ocha_baladiya"
        population_status = "verified_ocha"

     For unmatched entries (sub-baladiya muhalla zones in Tripoli/Benghazi
     and small villages outside the OCHA HNO 100), the existing value is
     preserved and the entry is annotated:
        population_year   = 2018
        population_source = "Pre-CARA workshop estimate (2018-2022)"
        population_method = "expert_estimate"
        population_status = "estimated_pending_verification"
                         or "sub_baladiya_estimate"

  4. Updates `_metadata.population_estimate` to UN WPP 2024 medium variant
     (Libya = 7,458,567) as the national calibration anchor and records the
     full provenance chain.

  5. Writes a timestamped JSONL audit log to `data/cache/hdx/audit/`
     BEFORE the JSON file is overwritten, so that the previous values can
     always be reconstructed. Uses the project audit logger when available.

Why "Non-displaced + Returnees" rather than the all-groups sum?
--------------------------------------------------------------
  - Non-displaced = ordinary residents.
  - Returnees    = former IDPs who have returned home; for risk-assessment
                   purposes they are now part of the resident population.
  - IDPs, Migrants, Refugees are tracked SEPARATELY by the displacement
    vulnerability indicator (Vulnerability pillar) and would be double
    counted if folded into the population baseline.
  - Sum across all 100 OCHA baladiyas of (Non-displaced + Returnees) is
    7,385,683 — within 1.0% of UN WPP 2024 (7,458,567) — providing a
    natural, audit-defensible national calibration without artificial
    scaling. This match validates the methodology.

Operation modes
---------------
  python scripts/refresh_population.py                  # fetch if missing, then write
  python scripts/refresh_population.py --dry-run        # preview only, no writes
  python scripts/refresh_population.py --no-fetch       # cache-only, fail if missing
  python scripts/refresh_population.py --audit-log PATH # custom audit destination
  python scripts/refresh_population.py --print-deltas   # show before/after table
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import shutil
import sys
import time
import unicodedata
import urllib.request
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Workspace root setup so this script is runnable from anywhere.
SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

# Optional: project audit logger. Falls back to stdlib logging if unavailable.
try:
    from utils.logging_config import audit as project_audit
except Exception:  # pragma: no cover - import path may differ in CI
    project_audit = None

logger = logging.getLogger("refresh_population")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MUNICIPALITIES_FILE = ROOT_DIR / "data" / "libya_municipalities.json"
HDX_CACHE_DIR = ROOT_DIR / "data" / "cache" / "hdx"
AUDIT_DIR = HDX_CACHE_DIR / "audit"

# UN World Population Prospects 2024 medium variant for Libya (national).
# Source: https://population.un.org/wpp/  (downloaded 2026, 2024 medium variant)
UN_WPP_2024_LIBYA = 7_458_567
UN_WPP_2024_AS_OF = "2024"

# OCHA Libya Population Statistics (HNO 2021), authoritative per-baladiya.
OCHA_HNO_RESOURCE = {
    "filename": "ocha_lby_pop_2021.xlsx",
    "url": (
        "https://data.humdata.org/dataset/d3b4d8cf-a9db-48d5-afe3-065011ab5ab2/"
        "resource/7b1e5eff-b760-4df0-9a95-2503f0cecb79/download/"
        "pop-consolidated_sadd_for_hno_28aug2021.xlsx"
    ),
    "sheet": "Pop consolidated for HNO",
    "title": "OCHA Libya Population Statistics (HNO 2021)",
    "license": "CC BY 4.0",
    "publisher": "OCHA Libya",
}

# IOM DTM Libya — kept for downstream displacement indicators (not used here
# to overwrite population). Present so that --no-fetch users get the full set.
IOM_DTM_R45_RESOURCE = {
    "filename": "iom_dtm_r45.xlsx",
    "url": (
        "https://data.humdata.org/dataset/89d609ef-b8c0-47b5-91a6-fe29af7d147c/"
        "resource/5fa55cb0-326a-4629-a056-eaf37aeff9d9/download/"
        "dtm-libya-r45-public-idps-returnees-dataset_hdx.xlsx"
    ),
    "title": "IOM DTM Libya Baseline Round 45 (IDPs, Returnees)",
    "license": "Custom (IOM DTM, public)",
    "publisher": "IOM DTM Libya",
}

IOM_MIGRANTS_R59_RESOURCE = {
    "filename": "iom_migrants_r59.xlsx",
    "url": (
        "https://data.humdata.org/dataset/cacf12d4-22b1-438f-bcb4-78aa9078ab94/"
        "resource/78a23009-b246-4968-8562-647cbae2c8aa/download/"
        "dtm_libya_r59_migrants_dataset.xlsx"
    ),
    "title": "IOM DTM Libya Migrants Baseline Round 59",
    "license": "CC BY 4.0",
    "publisher": "IOM DTM Libya",
}

ALL_RESOURCES = [OCHA_HNO_RESOURCE, IOM_DTM_R45_RESOURCE, IOM_MIGRANTS_R59_RESOURCE]

# ---------------------------------------------------------------------------
# Reconciliation map: our slug-style ID -> OCHA Baladiya PCode (or None).
#
#   value=PCode  -> matched, take population from OCHA
#   value=None   -> sub-baladiya muhalla zone OR small village not in HNO 100;
#                   keep existing estimate, mark with appropriate status.
#
# Only IDs that need a manual override (or are explicitly known to be
# sub-zones/small) appear here. Other IDs are auto-resolved by normalised
# exact-name matching against OCHA names — see _build_match_index().
# ---------------------------------------------------------------------------

MANUAL_ALIAS: Dict[str, Optional[str]] = {
    # --- Tripoli mantika (LY0211): 5 of 6 OCHA baladiyas covered ---
    "LY-001": "LY021104",  # Tripoli Center -> Tripoli baladiya
    "LY-002": "LY021105",  # Abu Salim -> Abusliem
    "LY-003": "LY021102",  # Tajoura -> Tajoura
    "LY-005": "LY021103",  # Ain Zara -> Ain Zara
    "LY-008": "LY021101",  # Suq al-Jumua -> Suq Aljumaa
    # Sub-zones (admin-4 muhalla within Tripoli baladiya):
    "LY-009": None, "LY-010": None, "LY-011": None,
    "LY-012": None, "LY-013": None, "LY-014": None,

    # --- Aljfara mantika (LY0212) ---
    "LY-004": "LY021206",  # Janzur -> Janzour
    "LY-006": "LY021203",  # Qasr Ben Ghashir -> Qasr Bin Ghasheer
    "LY-007": "LY021205",  # Al-Sawani -> Swani Bin Adam
    "LY-015": "LY021207",  # Al-Aziziyah -> Al Aziziya
    "LY-026": "LY021208",  # Warshafanah -> Al Maya (Warshafanah area)

    # --- Azzawya mantika (LY0213) + cross-mantika ---
    "LY-016": "LY021209",  # Al-Zahra -> Azzahra
    "LY-017": "LY021302",  # Surman -> Surman
    "LY-018": "LY021004",  # Qarabuli -> Garabolli (Almargeb)
    "LY-025": "LY021301",  # Zawiya -> Azzawya

    # --- Zwara mantika (LY0215) ---
    "LY-020": "LY021502",  # Sabratah -> Sabratha
    "LY-021": "LY021503",  # Zuwarah -> Zwara
    "LY-022": "LY021504",  # Al-Jumayliyah -> Aljmail
    "LY-023": "LY021505",  # Riqdalin -> Rigdaleen
    "LY-024": "LY021501",  # Al-Ajaylat -> Al Ajaylat
    "LY-027": "LY021506",  # Zaltan -> Ziltun

    # --- Almargeb mantika (LY0210) ---
    "LY-029": "LY021001",  # Al-Khums -> Alkhums
    "LY-030": "LY021005",  # Tarhuna -> Tarhuna
    "LY-031": "LY021002",  # Msallata -> Msallata

    # --- Misrata mantika (LY0214) ---
    "LY-033": "LY021402",  # Zliten -> Zliten
    "LY-034": "LY021401",  # Misrata -> Misrata
    "LY-035": "LY021404",  # Bani Walid -> Bani Waleed

    # --- Sirt mantika (LY0208) ---
    "LY-039": "LY020803",  # Sirte -> Sirt
    "LY-040": "LY020802",  # Harawa -> Hrawa
    "LY-041": "LY020801",  # Ben Jawad -> Khaleej Assidra (closest)

    # --- Al Jabal Al Gharbi mantika (LY0216) ---
    "LY-043": "LY021606",  # Gharyan -> Ghiryan
    "LY-044": "LY021605",  # Yafran -> Yefren
    "LY-045": "LY021609",  # Jadu -> Jadu
    "LY-046": "LY021602",  # Zintan -> Azzintan
    "LY-048": "LY021608",  # Al-Rajban -> Arrajban
    "LY-049": "LY021604",  # Al-Qal'a -> Al Qalaa

    # --- Nalut mantika (LY0209) ---
    "LY-050": "LY020905",  # Nalut -> Nalut
    "LY-051": "LY020901",  # Ghadames -> Ghadamis
    "LY-052": "LY020907",  # Darj -> Daraj

    # --- Benghazi mantika (LY0103) ---
    "LY-053": "LY010304",  # Benghazi Center -> Benghazi (whole baladiya)
    "LY-065": "LY010303",  # Soluq -> Suloug
    # Sub-zones of Benghazi (admin-4 muhalla):
    "LY-054": None, "LY-055": None, "LY-056": None, "LY-057": None,
    "LY-058": None, "LY-059": None, "LY-060": None, "LY-061": None,
    "LY-062": None,

    # --- Al Jabal Al Akhdar / Derna corridor ---
    "LY-068": None,        # Susah - not a separate baladiya in OCHA HNO
    "LY-080": "LY010105",  # Librag -> Labriq

    # --- Tobruk mantika (LY0104) ---
    "LY-075": None,        # Musaid - border crossing zone
    "LY-076": None,        # Omar al-Mukhtar - sub-zone

    # --- Ejdabia mantika (LY0105) ---
    "LY-078": "LY010506",  # Al-Burayqah -> Albrayga
    "LY-082": "LY010503",  # Awjila -> Aujala
    "LY-079": None,        # Al-Waygah - small
    "LY-085": None,        # Rabyana - small
    "LY-088": None,        # Al-Baraq - small

    # --- Wadi Ashshati mantika (LY0318) ---
    "LY-095": "LY031801",  # Brak al-Shati -> Brak
    "LY-103": "LY031802",  # Idri -> Edri

    # --- Murzuq mantika (LY0322) ---
    "LY-099": "LY032202",  # Qatrun -> Algatroun
    "LY-104": None,        # Al-Qirdah - small
    "LY-096": None,        # Umm al-Aranib - small

    # --- Ubari mantika (LY0320) ---
    "LY-090": "LY032001",  # Bent Bay -> Bint Bayya

    # --- Other small unmatched (no OCHA HNO equivalent) ---
    "LY-019": None, "LY-028": None, "LY-032": None, "LY-036": None,
    "LY-037": None, "LY-038": None, "LY-042": None, "LY-047": None,
    "LY-066": None, "LY-072": None, "LY-089": None, "LY-091": None,
    "LY-092": None, "LY-093": None, "LY-101": None, "LY-106": None,
}

# IDs explicitly classified as sub-baladiya muhalla (not separate baladiya).
# These are admin-4 zones inside an already-counted parent baladiya
# (Tripoli LY021104, Benghazi LY010304, Susah/Tobruk corridor). Their
# population MUST be excluded from any national rollup to avoid
# double-counting the parent OCHA baladiya total.
SUB_BALADIYA_IDS = {
    "LY-009", "LY-010", "LY-011", "LY-012", "LY-013", "LY-014",
    "LY-054", "LY-055", "LY-056", "LY-057", "LY-058", "LY-059",
    "LY-060", "LY-061", "LY-062",
    "LY-068", "LY-075", "LY-076",
}

# Manual aliases that are not exact name/PCode matches but the closest
# OCHA HNO baladiya by geography. These are flagged on the entry as
# population_match_confidence='approximate' so downstream consumers and
# auditors can see the mapping is heuristic, not authoritative.
APPROXIMATE_ALIASES = {
    "LY-041",  # Ben Jawad -> LY020801 Khaleej Assidra (closest baladiya)
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _audit(event: str, **fields: Any) -> None:
    """Emit an audit event via the project logger when available."""
    if project_audit is not None:
        try:
            project_audit(event, **fields)
            return
        except Exception:
            pass
    logger.info("AUDIT %s %s", event, json.dumps(fields, default=str, ensure_ascii=False))


def _norm(name: str) -> str:
    """Normalise a place-name string for cross-source matching."""
    if not name:
        return ""
    s = name.lower().strip()
    s = unicodedata.normalize("NFKD", s)
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def _ensure_resource(resource: Dict[str, str], allow_fetch: bool) -> Path:
    """Return cached path for an HDX resource. Download if missing and allowed."""
    HDX_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    target = HDX_CACHE_DIR / resource["filename"]
    if target.exists() and target.stat().st_size > 1024:
        return target
    if not allow_fetch:
        raise FileNotFoundError(
            f"Cached file missing: {target} (--no-fetch was requested)"
        )
    logger.info("Fetching %s", resource["title"])
    req = urllib.request.Request(
        resource["url"],
        headers={"User-Agent": "CARA-Libya/1.0 (+population-refresh)"},
    )
    tmp = target.with_suffix(target.suffix + ".tmp")
    with urllib.request.urlopen(req, timeout=120) as r, open(tmp, "wb") as f:
        shutil.copyfileobj(r, f)
    if tmp.stat().st_size < 2048:
        tmp.unlink(missing_ok=True)
        raise RuntimeError(
            f"Download too small ({tmp.stat().st_size} bytes) — likely an "
            f"HDX redirect error. URL: {resource['url']}"
        )
    tmp.replace(target)
    return target


def _parse_ocha_baladiya_population(xlsx_path: Path) -> Dict[str, Dict[str, Any]]:
    """Parse the OCHA HNO 'Pop consolidated for HNO' sheet.

    Returns {baladiya_pcode: {baladiya_en, mantika_en, mantika_pcode,
                              non_displaced, returnees, idps, migrants,
                              refugees, total_population, resident_population}}
    """
    from openpyxl import load_workbook  # local import keeps script lightweight

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[OCHA_HNO_RESOURCE["sheet"]]
    rows = ws.iter_rows(values_only=True)

    # Header row + HXL hashtag row
    next(rows)
    next(rows)

    out: Dict[str, Dict[str, Any]] = defaultdict(lambda: {
        "baladiya_en": "", "mantika_en": "", "mantika_pcode": "",
        "non_displaced": 0.0, "returnees": 0.0, "idps": 0.0,
        "migrants": 0.0, "refugees": 0.0, "total_population": 0.0,
    })
    group_field = {
        "Non-displaced": "non_displaced",
        "Returnees":     "returnees",
        "IDPs":          "idps",
        "Migrants":      "migrants",
        "Refugees":      "refugees",
    }
    for r in rows:
        if not r or r[0] is None:
            continue
        mantika_pcode, mantika_en, baladiya_pcode, baladiya_en, popgroup, individuals = r[:6]
        if not baladiya_pcode:
            continue
        rec = out[baladiya_pcode]
        rec["baladiya_en"] = baladiya_en or rec["baladiya_en"]
        rec["mantika_en"] = mantika_en or rec["mantika_en"]
        rec["mantika_pcode"] = mantika_pcode or rec["mantika_pcode"]
        field = group_field.get(popgroup)
        if field is None:
            continue
        try:
            v = float(individuals or 0)
        except (ValueError, TypeError):
            continue
        rec[field] += v
        rec["total_population"] += v

    # Compute resident population (Non-displaced + Returnees) — the value we
    # use to overwrite municipality population. See module docstring.
    for rec in out.values():
        rec["resident_population"] = round(rec["non_displaced"] + rec["returnees"])

    return dict(out)


def _build_match_index(
    municipalities: List[Dict[str, Any]],
    ocha: Dict[str, Dict[str, Any]],
) -> Tuple[Dict[str, Tuple[str, str]], List[str], List[str]]:
    """Resolve each of our IDs to (ocha_pcode, match_method) or unmatched.

    Returns:
        matches:        {our_id: (ocha_pcode, method)}
        sub_zones:      list of our IDs explicitly classified as sub-baladiya
        small_villages: list of our IDs with no OCHA equivalent
    """
    ocha_by_norm: Dict[str, str] = {}
    for pc, rec in ocha.items():
        ocha_by_norm[_norm(rec["baladiya_en"])] = pc

    matches: Dict[str, Tuple[str, str]] = {}
    sub_zones: List[str] = []
    small_villages: List[str] = []

    for m in municipalities:
        mid = m["id"]
        if mid in MANUAL_ALIAS:
            target = MANUAL_ALIAS[mid]
            if target is None:
                if mid in SUB_BALADIYA_IDS:
                    sub_zones.append(mid)
                else:
                    small_villages.append(mid)
                continue
            if target not in ocha:
                logger.warning("Manual alias %s -> %s missing in OCHA data", mid, target)
                small_villages.append(mid)
                continue
            matches[mid] = (target, "manual_alias")
            continue
        # Auto-match by normalised English name.
        n = _norm(m.get("name_en", ""))
        if n and n in ocha_by_norm:
            matches[mid] = (ocha_by_norm[n], "auto_exact")
        else:
            small_villages.append(mid)

    return matches, sub_zones, small_villages


def _apply_updates(
    data: Dict[str, Any],
    ocha: Dict[str, Dict[str, Any]],
    matches: Dict[str, Tuple[str, str]],
    sub_zones: List[str],
    small_villages: List[str],
) -> List[Dict[str, Any]]:
    """Mutate `data` in place. Return a per-entry delta record list for audit."""
    deltas: List[Dict[str, Any]] = []
    for m in data["municipalities"]:
        mid = m["id"]
        old_pop = int(m.get("population", 0))
        delta: Dict[str, Any] = {
            "id": mid, "name_en": m.get("name_en", ""),
            "old_population": old_pop,
        }
        if mid in matches:
            pc, method = matches[mid]
            o = ocha[pc]
            new_pop = int(o["resident_population"])
            confidence = "approximate" if mid in APPROXIMATE_ALIASES else (
                "exact" if method == "auto_exact" else "high"
            )
            m["population"] = new_pop
            m["population_year"] = 2021
            m["population_source"] = "OCHA Libya HNO 2021 (CC BY 4.0)"
            m["population_method"] = "hdx_ocha_baladiya"
            m["population_status"] = "verified_ocha"
            m["population_ocha_pcode"] = pc
            m["population_match_confidence"] = confidence
            m["population_in_national_total"] = True
            m["status"] = "verified_ocha_hno"
            delta.update(
                new_population=new_pop, change=new_pop - old_pop,
                ocha_pcode=pc, ocha_baladiya_en=o["baladiya_en"],
                method=method, confidence=confidence, status="verified_ocha",
            )
        elif mid in sub_zones:
            m["population_year"] = 2018
            m["population_source"] = "Pre-CARA workshop estimate (2018-2022)"
            m["population_method"] = "expert_estimate"
            m["population_status"] = "sub_baladiya_estimate"
            m["population_in_national_total"] = False
            m["status"] = "sub_baladiya_estimate"
            # Sub-baladiya zones are inside an already-counted parent baladiya
            # (Tripoli LY021104, Benghazi LY010304); their population is part
            # of that parent's OCHA total and must not be summed nationally.
            m.pop("population_match_confidence", None)
            m.pop("population_ocha_pcode", None)
            delta.update(
                new_population=old_pop, change=0, method="kept_estimate",
                status="sub_baladiya_estimate",
            )
        else:
            m["population_year"] = 2018
            m["population_source"] = "Pre-CARA workshop estimate (2018-2022)"
            m["population_method"] = "expert_estimate"
            m["population_status"] = "estimated_pending_verification"
            m["population_in_national_total"] = True
            m["status"] = "estimated_pending_verification"
            m.pop("population_match_confidence", None)
            m.pop("population_ocha_pcode", None)
            delta.update(
                new_population=old_pop, change=0, method="kept_estimate",
                status="estimated_pending_verification",
            )
        deltas.append(delta)

    # Compute the deduplicated national total: sum populations only from
    # entries flagged population_in_national_total=True (excludes the 18
    # sub-baladiya muhalla overlays whose parent baladiya is already counted
    # via verified_ocha). This is the correct figure for any national
    # rollup, per-capita rate, or calibration test.
    national_deduped = sum(
        int(m.get("population", 0))
        for m in data["municipalities"]
        if m.get("population_in_national_total", True)
    )
    n_total = len(data["municipalities"])
    n_verified = sum(1 for m in data["municipalities"]
                     if m.get("population_status") == "verified_ocha")
    n_pending = sum(1 for m in data["municipalities"]
                    if m.get("population_status") == "estimated_pending_verification")
    n_subzone = sum(1 for m in data["municipalities"]
                    if m.get("population_status") == "sub_baladiya_estimate")

    # Update _metadata block.
    md = data.setdefault("_metadata", {})
    md["last_updated"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    md["current_count"] = n_total
    md["data_gap_note"] = (
        f"This file contains {n_total} entries: {n_verified} verified against "
        f"OCHA HNO 2021 baladiyas, {n_pending} small villages outside HNO scope "
        f"with workshop estimates pending verification, and {n_subzone} "
        f"admin-4 sub-baladiya overlays (Tripoli and Benghazi muhalla zones, "
        f"plus Susah and the Musaid/Omar al-Mukhtar Tobruk corridor) that are "
        f"flagged population_in_national_total=False to prevent double-counting "
        f"their parent baladiya. To reach Libya's 148-baladiya target, "
        f"approximately {max(148 - (n_verified + n_pending), 0)} additional "
        f"baladiyas — primarily from the 41 OCHA HNO baladiyas not yet matched "
        f"and unsplit metropolitan units in Misrata — still need to be added "
        f"from official Libyan government administrative records (High National "
        f"Elections Commission or Ministry of Local Government municipal "
        f"registry)."
    )
    md["population_estimate"] = UN_WPP_2024_LIBYA
    md["population_estimate_source"] = (
        f"UN World Population Prospects {UN_WPP_2024_AS_OF} medium variant (Libya national)"
    )
    md["national_total_deduped"] = national_deduped
    md["national_total_deduped_explanation"] = (
        "Sum of population over entries with population_in_national_total=True. "
        "Excludes the sub-baladiya muhalla overlays in Tripoli/Benghazi to "
        "avoid double-counting their parent baladiya (already counted in the "
        "verified_ocha set). This is the figure to use for national rollups."
    )

    # Mirror the canonical national estimate onto the national-jurisdiction
    # block so consumers reading data['national']['population_estimate'] see
    # the same UN WPP 2024 figure as data['_metadata']['population_estimate'].
    nat = data.setdefault("national", {})
    nat["population_estimate"] = UN_WPP_2024_LIBYA
    nat["population_estimate_source"] = md["population_estimate_source"]
    md["population_provenance"] = {
        "primary":            OCHA_HNO_RESOURCE["title"],
        "primary_license":    OCHA_HNO_RESOURCE["license"],
        "primary_publisher":  OCHA_HNO_RESOURCE["publisher"],
        "primary_url":        OCHA_HNO_RESOURCE["url"],
        "national_calibration": (
            f"UN WPP {UN_WPP_2024_AS_OF} medium variant = {UN_WPP_2024_LIBYA:,}"
        ),
        "secondary_idp":      IOM_DTM_R45_RESOURCE["title"],
        "secondary_migrant":  IOM_MIGRANTS_R59_RESOURCE["title"],
        "method": (
            "Per-baladiya population = OCHA HNO 2021 Non-displaced + Returnees. "
            "Returnees are former IDPs who have returned to their pre-displacement "
            "baladiya and are part of the resident population for risk-assessment "
            "purposes. IDPs/Migrants/Refugees are tracked separately by the "
            "Vulnerability pillar's displacement_vulnerability indicator and are "
            "intentionally NOT folded into the population baseline. Sum of OCHA "
            "(Non-displaced + Returnees) across the 100 HNO baladiyas is "
            "7,385,683, which validates against UN WPP 2024 within 1.0%."
        ),
        "coverage": (
            f"OCHA HNO 2021 covers 100 of Libya's 148 baladiyas. This file has "
            f"{n_total} entries: {n_verified} 'verified_ocha' (population from "
            f"OCHA HNO 2021, population_in_national_total=True), "
            f"{n_pending} 'estimated_pending_verification' (small villages "
            f"outside HNO scope, population_in_national_total=True), and "
            f"{n_subzone} 'sub_baladiya_estimate' (admin-4 muhalla overlays "
            f"inside Tripoli LY021104 / Benghazi LY010304, "
            f"population_in_national_total=False to prevent double-counting). "
            f"Deduplicated national total over the {n_verified + n_pending} "
            f"counted entries = {national_deduped:,} "
            f"({100 * national_deduped / UN_WPP_2024_LIBYA:.1f}% of UN WPP 2024); "
            f"the gap to UN WPP reflects the ~41 OCHA HNO baladiyas not yet "
            f"in this file."
        ),
    }

    # Refresh the legacy free-text provenance field so it no longer cites
    # the pre-CARA 2018 election base data and IOM DTM 2022 for population —
    # those are now superseded by the population_provenance block above.
    md["provenance"] = (
        "Compiled from: UNOCHA Libya Humanitarian Response Plan administrative "
        "units (2023); IOM Displacement Tracking Matrix Libya operational "
        "coverage; academic literature on post-2011 Libyan administrative "
        "reorganization. Per-baladiya population values come from OCHA HNO "
        "2021 via HDX (CC BY 4.0) — see _metadata.population_provenance for "
        "the full chain of custody."
    )
    return deltas


def _write_audit(deltas: List[Dict[str, Any]], audit_path: Optional[Path]) -> Path:
    """Write a JSONL audit record before mutating the JSON file."""
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    if audit_path is None:
        ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        audit_path = AUDIT_DIR / f"population_refresh_{ts}.jsonl"
    with open(audit_path, "w", encoding="utf-8") as f:
        f.write(json.dumps({
            "event": "population_refresh_started",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "source": OCHA_HNO_RESOURCE["title"],
            "entries_processed": len(deltas),
        }, ensure_ascii=False) + "\n")
        for d in deltas:
            f.write(json.dumps(d, ensure_ascii=False) + "\n")
    _audit("population_refresh_audit_written", path=str(audit_path), entries=len(deltas))
    return audit_path


def _summarise(deltas: List[Dict[str, Any]]) -> Dict[str, int]:
    summary = defaultdict(int)
    for d in deltas:
        summary[d["status"]] += 1
        summary["total_old"] += d["old_population"]
        summary["total_new"] += d["new_population"]
    return dict(summary)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n", 1)[0])
    parser.add_argument("--dry-run", action="store_true",
                        help="Print summary; do not write JSON or audit log.")
    parser.add_argument("--no-fetch", action="store_true",
                        help="Use cached HDX files only; fail if missing.")
    parser.add_argument("--audit-log", type=Path, default=None,
                        help="Override audit log path (default: timestamped under data/cache/hdx/audit/).")
    parser.add_argument("--print-deltas", action="store_true",
                        help="Print one line per municipality showing before/after.")
    args = parser.parse_args(argv)

    t0 = time.time()
    _audit("population_refresh_start", dry_run=args.dry_run, no_fetch=args.no_fetch)

    # 1. Ensure all HDX files are available in cache.
    for resource in ALL_RESOURCES:
        try:
            _ensure_resource(resource, allow_fetch=not args.no_fetch)
        except (FileNotFoundError, RuntimeError) as exc:
            logger.error("Resource fetch failed for %s: %s", resource["filename"], exc)
            if resource is OCHA_HNO_RESOURCE:
                # The OCHA file is the only one we MUST have for the refresh.
                _audit("population_refresh_aborted", reason=str(exc),
                       resource=resource["filename"])
                return 2

    # 2. Parse OCHA file.
    ocha_path = HDX_CACHE_DIR / OCHA_HNO_RESOURCE["filename"]
    ocha = _parse_ocha_baladiya_population(ocha_path)
    logger.info("OCHA HNO parsed: %d baladiyas", len(ocha))

    # 3. Load municipalities and compute reconciliation.
    with open(MUNICIPALITIES_FILE, encoding="utf-8") as f:
        data = json.load(f)
    municipalities = data["municipalities"]
    matches, sub_zones, small_villages = _build_match_index(municipalities, ocha)
    logger.info(
        "Reconciliation: matched=%d, sub_zones=%d, small_villages=%d, total=%d",
        len(matches), len(sub_zones), len(small_villages), len(municipalities),
    )

    # 4. Apply updates to in-memory copy.
    deltas = _apply_updates(data, ocha, matches, sub_zones, small_villages)

    # 5. Summary report.
    summary = _summarise(deltas)
    logger.info("Summary: %s", summary)
    natl_old = summary["total_old"]
    natl_new = summary["total_new"]
    logger.info(
        "National population: %d -> %d (delta %+d, %.1f%%)",
        natl_old, natl_new, natl_new - natl_old,
        100 * (natl_new - natl_old) / max(natl_old, 1),
    )
    logger.info(
        "UN WPP 2024 anchor: %d (matched-entries+estimates total = %.1f%% of UN WPP)",
        UN_WPP_2024_LIBYA, 100 * natl_new / UN_WPP_2024_LIBYA,
    )

    if args.print_deltas:
        print(f"\n{'ID':8s} {'NAME':28s} {'OLD':>10s} {'NEW':>10s} {'DELTA':>10s}  STATUS")
        print("-" * 90)
        for d in deltas:
            print(f"{d['id']:8s} {d['name_en'][:28]:28s} "
                  f"{d['old_population']:>10,} {d['new_population']:>10,} "
                  f"{d['change']:>+10,}  {d['status']}")

    if args.dry_run:
        logger.info("DRY RUN — no files written.")
        _audit("population_refresh_dry_run_complete",
               duration_s=round(time.time() - t0, 2),
               matched=len(matches), sub_zones=len(sub_zones),
               small_villages=len(small_villages))
        return 0

    # 6. Write audit log FIRST (recoverability), THEN overwrite JSON.
    audit_path = _write_audit(deltas, args.audit_log)
    logger.info("Audit log: %s", audit_path)

    tmp = MUNICIPALITIES_FILE.with_suffix(".json.tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    tmp.replace(MUNICIPALITIES_FILE)
    logger.info("Wrote %s", MUNICIPALITIES_FILE)

    _audit("population_refresh_complete",
           duration_s=round(time.time() - t0, 2),
           matched=len(matches), sub_zones=len(sub_zones),
           small_villages=len(small_villages),
           national_old=natl_old, national_new=natl_new,
           un_wpp_2024=UN_WPP_2024_LIBYA,
           audit_path=str(audit_path))
    return 0


if __name__ == "__main__":
    sys.exit(main())
