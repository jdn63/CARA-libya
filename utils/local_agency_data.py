"""Local-agency data ingestion engine — domain-driven.

This module is **schema-agnostic**: it consumes a :class:`DomainSpec` from
:mod:`utils.data_entry_domains` and produces

1. the downloadable Excel template that local response agencies fill in
   (one row per Libyan municipality, bilingual headers, RTL, validations,
   per-domain Instructions sheet);
2. a consolidator that reads every workbook in the per-domain upload
   directory and returns the most-recent capture per municipality
   (deterministic tie-break by upload mtime so re-submissions on the same
   day overwrite older entries from earlier files);
3. an export workbook that mirrors the comparison table back to .xlsx, with
   spreadsheet-formula injection neutralised on every text cell.

All five workshop domains (infectious disease, vector-borne disease, NCDs,
MNCH, environmental health) — and any future ones added to ``DOMAINS`` —
are handled by this single pipeline.
"""

from __future__ import annotations

import io
import json
import logging
import os
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.packaging.custom import CustomPropertyList, StringProperty
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from utils.data_entry_domains import DomainSpec, Indicator

# Custom workbook property used to identify which domain a template / upload
# belongs to. The ``domain_upload`` route checks this on every upload so that
# the infectious-disease template cannot be silently mis-routed to MNCH (or
# any other domain whose column count happens to match).
DOMAIN_PROPERTY_NAME = "cara_libya_domain_key"

# Marker value stored in DOMAIN_PROPERTY_NAME for the combined Master
# Template workbook. Distinct from any real domain key so a master workbook
# posted to a single-domain endpoint is rejected and vice-versa.
MASTER_DOMAIN_KEY = "__libya_cara_master__"

# Earliest date we accept on the server side. Excel-side validation enforces
# the same bound, but we mirror it here so a user who bypasses the workbook
# validation cannot poison the consolidated table with year-1900 rows.
EARLIEST_CAPTURE_DATE = date(2000, 1, 1)

logger = logging.getLogger(__name__)

# --------------------------------------------------------------------------- #
# Constants
# --------------------------------------------------------------------------- #

UPLOAD_ROOT = Path("data/uploads/local_agencies")
MUNICIPALITIES_FILE = Path("data/libya_municipalities.json")

# Fixed columns that come before the indicator block, in order.
FIXED_COLUMNS: list[tuple[str, str, str]] = [
    ("municipality_id", "رمز البلدية",            "Municipality ID"),
    ("name_ar",         "اسم البلدية (عربي)",     "Municipality (Arabic)"),
    ("name_en",         "اسم البلدية (إنجليزي)",  "Municipality (English)"),
    ("capture_date",    "تاريخ جمع البيانات",     "Date of Data Capture (YYYY-MM-DD)"),
    ("agency_name",     "اسم الجهة المُبلِّغة",    "Reporting Agency"),
]

NOTES_COLUMN: tuple[str, str, str] = (
    "notes", "ملاحظات / تباينات في البيانات",
    "Notes / Data Discrepancies",
)


def _indicator_header(ind: Indicator) -> tuple[str, str]:
    """Build the bilingual two-line header for an indicator column.

    The header always carries the unit so reporting agencies do not have to
    refer back to a separate codebook while filling the template.
    """
    if ind.group_ar:
        ar = f"{ind.group_ar} — {ind.name_ar} ({ind.unit_ar})"
        en = f"{ind.group_en} — {ind.name_en} ({ind.unit_en})"
    else:
        ar = f"{ind.name_ar} ({ind.unit_ar})"
        en = f"{ind.name_en} ({ind.unit_en})"
    return ar, en


def all_columns(spec: DomainSpec) -> list[tuple[str, str, str]]:
    """Full ordered list of (key, ar_label, en_label) tuples for a domain."""
    cols: list[tuple[str, str, str]] = list(FIXED_COLUMNS)
    for ind in spec.indicators:
        ar, en = _indicator_header(ind)
        cols.append((ind.code, ar, en))
    cols.append(NOTES_COLUMN)
    return cols


def upload_dir_for(spec: DomainSpec) -> Path:
    """Per-domain upload directory under ``data/uploads/local_agencies/``."""
    safe = spec.key.replace("-", "_")
    return UPLOAD_ROOT / safe


# --------------------------------------------------------------------------- #
# Municipality loader
# --------------------------------------------------------------------------- #

def _load_municipalities() -> list[dict[str, Any]]:
    try:
        with MUNICIPALITIES_FILE.open(encoding="utf-8") as f:
            doc = json.load(f)
        return doc.get("municipalities", [])
    except FileNotFoundError:
        logger.warning("Municipality file %s not found", MUNICIPALITIES_FILE)
        return []
    except json.JSONDecodeError as exc:
        logger.error("Municipality file is not valid JSON: %s", exc)
        return []


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def _safe_text(value: Any) -> str:
    """Neutralise spreadsheet formula injection on export.

    Cells whose first character is ``= + - @`` (or a leading tab/CR that
    some clients strip) are interpreted as formulas by Excel/LibreOffice
    when re-opened. Prefix such values with a single quote so they render
    as literal text.
    """
    if value is None:
        return ""
    s = str(value)
    if s and s[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + s
    return s


def _styled_header(cell, ar_label: str, en_label: str) -> None:
    cell.value = f"{ar_label}\n{en_label}"
    cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1f4e79")
    cell.alignment = Alignment(
        wrap_text=True, vertical="center", horizontal="center"
    )
    thin = Side(border_style="thin", color="BFBFBF")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# --------------------------------------------------------------------------- #
# Template builder
# --------------------------------------------------------------------------- #

def _populate_data_entry_sheet(ws, spec: DomainSpec,
                               prefill_municipalities: bool = True) -> None:
    """Populate ``ws`` as a Data-Entry sheet for ``spec``.

    Shared by both the single-domain template and the Master Template
    (which calls this once per tab).
    """
    ws.sheet_view.rightToLeft = True
    columns = all_columns(spec)

    # Header row.
    for col_idx, (_key, ar, en) in enumerate(columns, start=1):
        _styled_header(ws.cell(row=1, column=col_idx), ar, en)

    # Pre-populate municipality rows.
    municipalities = _load_municipalities() if prefill_municipalities else []
    for row_idx, m in enumerate(municipalities, start=2):
        ws.cell(row=row_idx, column=1, value=_safe_text(m.get("id", "")))
        ws.cell(row=row_idx, column=2, value=_safe_text(m.get("name_ar", "")))
        ws.cell(row=row_idx, column=3, value=_safe_text(m.get("name_en", "")))

    # Column widths.
    width_map = {
        "municipality_id": 14, "name_ar": 26, "name_en": 26,
        "capture_date": 18, "agency_name": 28, "notes": 40,
    }
    for col_idx, (key, _ar, _en) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            width_map.get(key, 22)
        )

    ws.freeze_panes = "F2"
    ws.row_dimensions[1].height = 56

    # Date validation on capture-date column.
    capture_col_idx = next(
        i for i, (k, *_rest) in enumerate(columns, start=1)
        if k == "capture_date"
    )
    capture_letter = get_column_letter(capture_col_idx)
    last_row = max(2 + len(municipalities), 200)
    date_dv = DataValidation(
        type="date", operator="greaterThan", formula1="2000-01-01",
        allow_blank=True, showErrorMessage=True,
        errorTitle="Invalid date",
        error="Use a real date on or after 2000-01-01 (format YYYY-MM-DD).",
    )
    date_dv.add(f"{capture_letter}2:{capture_letter}{last_row}")
    ws.add_data_validation(date_dv)

    # Numeric validation per indicator.
    indicator_cols: dict[str, int] = {
        key: idx for idx, (key, *_r) in enumerate(columns, start=1)
        if key not in {"municipality_id", "name_ar", "name_en",
                       "capture_date", "agency_name", "notes"}
    }
    for ind in spec.indicators:
        col_idx = indicator_cols[ind.code]
        letter = get_column_letter(col_idx)
        if ind.max_value is None:
            dv = DataValidation(
                type="decimal", operator="greaterThanOrEqual",
                formula1=ind.min_value, allow_blank=True,
                showErrorMessage=True, errorTitle="Invalid value",
                error=f"Value must be >= {ind.min_value}.",
            )
        else:
            dv = DataValidation(
                type="decimal", operator="between",
                formula1=ind.min_value, formula2=ind.max_value,
                allow_blank=True, showErrorMessage=True,
                errorTitle="Invalid value",
                error=(f"Value must be between {ind.min_value} "
                       f"and {ind.max_value}."),
            )
        dv.add(f"{letter}2:{letter}{last_row}")
        ws.add_data_validation(dv)


def _set_domain_property(wb, value: str) -> None:
    """Attach the ``cara_libya_domain_key`` custom workbook property."""
    props = CustomPropertyList()
    props.append(StringProperty(name=DOMAIN_PROPERTY_NAME, value=value))
    wb.custom_doc_props = props


def build_template_workbook(spec: DomainSpec) -> bytes:
    """Build the single-domain Excel template that agencies download.

    Returns
    -------
    bytes
        Raw .xlsx bytes ready for Flask's ``send_file``.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Entry"
    _populate_data_entry_sheet(ws, spec, prefill_municipalities=True)
    _set_domain_property(wb, spec.key)

    # Instructions sheet — bilingual, RTL.
    inst = wb.create_sheet("Instructions")
    inst.sheet_view.rightToLeft = True
    inst.column_dimensions["A"].width = 110
    instructions: list[tuple[str, bool]] = [
        (f"{spec.name_ar} — التعليمات", True),
        (f"{spec.name_en} — Instructions", True),
        ("", False),
        ("١. لا تُعدِّل صف الترويسة أو ترتيب الأعمدة.", False),
        ("1. Do not modify the header row or the column order.", False),
        ("", False),
        ("٢. تاريخ جمع البيانات بصيغة YYYY-MM-DD (مثال: 2026-04-23).", False),
        ("2. Date of Data Capture must follow the format YYYY-MM-DD "
         "(e.g. 2026-04-23).", False),
        ("", False),
        ("٣. وحدة كل عمود مذكورة بين قوسين في الترويسة. اترك الخلية فارغة "
         "إذا لم تتوفر البيانات — لا تكتب صفراً ولا «N/A».", False),
        ("3. The unit for each column is shown in parentheses in the "
         "header. Leave the cell empty if data is unavailable — do NOT "
         "enter zero or 'N/A'.", False),
        ("", False),
        ("٤. أَدخِل صفّاً واحداً لكل بلدية لكل تاريخ جمع. لإرسال تحديث "
         "أحدث، أَضِف صفّاً جديداً — لا تَستبدل الصف القديم.", False),
        ("4. Enter one row per municipality per capture date. To submit "
         "an update later, add a new row — do not overwrite the old row.",
         False),
        ("", False),
        ("٥. استخدم حقل «الملاحظات» للإشارة إلى أيّ تباين في البيانات أو "
         "ظرف خاص (نقص في الإبلاغ، تفشٍّ موضعي، تغيير منهجي…).", False),
        ("5. Use the Notes field to flag any data discrepancies or "
         "special circumstances (under-reporting, localised outbreak, "
         "methodology change, etc.).", False),
        ("", False),
        ("٦. عند الانتهاء، احفظ الملف بصيغة .xlsx وارفعه إلى البوابة.",
         False),
        ("6. When done, save the file as .xlsx and upload it through the "
         "portal.", False),
        ("", False),
        (f"دور هذا المجال في INFORM: {spec.inform_role_ar}", False),
        (f"INFORM linkage: {spec.inform_role_en}", False),
    ]
    for i, (text, is_heading) in enumerate(instructions, start=1):
        cell = inst.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if is_heading:
            cell.font = Font(bold=True, size=14, color="1f4e79")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Master template (all domains in one workbook)
# --------------------------------------------------------------------------- #

def build_master_template_workbook(
    domains: tuple[DomainSpec, ...]
) -> bytes:
    """Build the combined Libya CARA Master Template with one tab per domain.

    The first tab is a bilingual Instructions sheet. Each subsequent tab
    carries one domain's full data-entry grid (prefilled with the 148
    municipality IDs and names, with per-column data validation).

    The workbook-level ``cara_libya_domain_key`` property is set to
    :data:`MASTER_DOMAIN_KEY` so that single-domain upload endpoints reject
    it and vice-versa.
    """
    wb = Workbook()
    # Repurpose the default sheet as Instructions.
    inst = wb.active
    inst.title = "Instructions"
    inst.sheet_view.rightToLeft = True
    inst.column_dimensions["A"].width = 110

    lines: list[tuple[str, bool]] = [
        ("قالب ليبيا CARA الموحَّد — تعليمات", True),
        ("Libya CARA Master Template — Instructions", True),
        ("", False),
        ("يحتوي هذا الملف على تبويب منفصل لكل مجال من مجالات الاستجابة "
         "الخمسة. يمكنك تعبئة ما توفّر لديك فقط ورفع الملف كما هو — "
         "التبويبات الفارغة ستُتجاهل تلقائياً ولن تمسح البيانات السابقة.",
         False),
        ("This workbook has one tab per response domain. Fill in only the "
         "tabs for which you have data and upload the file — any tab that "
         "is left empty will be skipped and will NOT overwrite previously "
         "submitted data.",
         False),
        ("", False),
        ("المجالات المتوفّرة في هذا الملف / Tabs in this workbook:", True),
    ]
    for spec in domains:
        lines.append((
            f"   •  «{spec.resolved_sheet_title()}»   —   "
            f"{spec.name_ar}   /   {spec.name_en}",
            False,
        ))
    lines += [
        ("", False),
        ("قواعد عامة / General rules", True),
        ("١. لا تُعدِّل صفّ الترويسة ولا ترتيب الأعمدة في أي تبويب.", False),
        ("1. Do not modify the header row or column order in any tab.",
         False),
        ("", False),
        ("٢. تاريخ جمع البيانات بصيغة YYYY-MM-DD (مثال: 2026-04-23). "
         "الوحدة لكل عمود مذكورة بين قوسين في ترويسته.", False),
        ("2. Date of Data Capture format is YYYY-MM-DD. Each column's "
         "unit is shown in parentheses in its header.", False),
        ("", False),
        ("٣. اترك الخلية فارغة إذا لم تتوفر البيانات — لا تكتب «0» ولا "
         "«N/A». صف يحمل رمز بلدية وتاريخاً دون أي قيمة مؤشر سيُتجاهل.",
         False),
        ("3. Leave a cell empty if a value is unavailable — do NOT enter "
         "0 or 'N/A'. A row with a municipality ID and a date but no "
         "indicator values will be ignored.", False),
        ("", False),
        ("٤. لإرسال تحديث لاحقاً، أَضِف صفّاً جديداً بتاريخ جديد في نفس "
         "التبويب — لا تَستبدل الصف القديم.", False),
        ("4. To submit an update later, add a new row with a new capture "
         "date in the same tab — do not overwrite the old row.", False),
        ("", False),
        ("٥. عند الانتهاء احفظ الملف بصيغة .xlsx وارفعه عبر صفحة «القالب "
         "الموحَّد».", False),
        ("5. When done, save as .xlsx and upload via the Master Template "
         "page of the portal.", False),
    ]
    for i, (text, is_heading) in enumerate(lines, start=1):
        cell = inst.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if is_heading:
            cell.font = Font(bold=True, size=13, color="1f4e79")

    # One Data-Entry tab per domain.
    for spec in domains:
        ws = wb.create_sheet(title=spec.resolved_sheet_title())
        _populate_data_entry_sheet(ws, spec, prefill_municipalities=True)

    _set_domain_property(wb, MASTER_DOMAIN_KEY)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def split_master_upload(
    raw: bytes, domains: tuple[DomainSpec, ...]
) -> tuple[dict[str, bytes], list[str]]:
    """Split an uploaded Master workbook into per-domain single-sheet files.

    Returns
    -------
    (accepted, skipped_empty)
        * ``accepted`` is a ``{domain_key: bytes}`` mapping — each value is
          a standalone single-domain workbook (with the correct per-domain
          marker) ready to be written into the domain's upload directory.
          Only domains whose tab contains at least one row with a
          municipality ID, a valid capture date, AND at least one non-blank
          indicator value are included.
        * ``skipped_empty`` lists the domain keys whose tab was either
          missing or contained no usable data rows.
    """
    try:
        src = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover — guarded upstream
        logger.warning("Could not open master workbook: %s", exc)
        return {}, [s.key for s in domains]

    accepted: dict[str, bytes] = {}
    skipped: list[str] = []

    for spec in domains:
        title = spec.resolved_sheet_title()
        if title not in src.sheetnames:
            skipped.append(spec.key)
            continue
        src_ws = src[title]
        columns = all_columns(spec)
        key_idx = {k: i for i, (k, *_r) in enumerate(columns)}
        data_rows: list[list[Any]] = []
        for excel_row in src_ws.iter_rows(min_row=2, values_only=True):
            if excel_row is None:
                continue
            padded = list(excel_row) + [None] * (len(columns) - len(excel_row))
            muni_id = padded[key_idx["municipality_id"]]
            # The row counts as a real submission only when the date parses
            # and at least one indicator parses as a valid number after
            # bounds checks. Anything short of that will be dropped by
            # _read_workbook anyway — counting it now would cause us to
            # report "imported" for tabs that actually ingest nothing.
            if not muni_id:
                continue
            if _parse_date(padded[key_idx["capture_date"]]) is None:
                continue
            has_valid_indicator = any(
                _parse_value(padded[key_idx[ind.code]], ind) is not None
                for ind in spec.indicators
            )
            if not has_valid_indicator:
                continue
            data_rows.append(padded)

        if not data_rows:
            skipped.append(spec.key)
            continue

        # Build a clean single-domain workbook containing exactly the rows
        # this agency actually submitted. Writing it through the usual
        # template scaffolding ensures _read_workbook consumes it exactly
        # the way it would a per-domain upload.
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "Data Entry"
        _populate_data_entry_sheet(out_ws, spec, prefill_municipalities=False)
        for r_idx, padded in enumerate(data_rows, start=2):
            for c_idx, val in enumerate(padded, start=1):
                out_ws.cell(row=r_idx, column=c_idx, value=val)
        _set_domain_property(out_wb, spec.key)
        buf = io.BytesIO()
        out_wb.save(buf)
        accepted[spec.key] = buf.getvalue()

    return accepted, skipped


def build_master_export_workbook(
    domains: tuple[DomainSpec, ...]
) -> bytes:
    """Build a combined comparison workbook — one tab per domain, each
    containing the same consolidated table the single-domain export uses.
    """
    wb = Workbook()
    wb.remove(wb.active)  # we'll add our own tabs
    for spec in domains:
        table = consolidated_table(spec)
        columns = all_columns(spec)
        ws = wb.create_sheet(title=spec.resolved_sheet_title())
        ws.sheet_view.rightToLeft = True

        freshness = table.get("freshness")
        fresh_text = (
            f"{spec.name_ar}   |   {spec.name_en}   ||   "
            f"تاريخ أحدث البيانات: "
            f"{freshness.isoformat() if freshness else 'غير متاح'}   |   "
            f"Latest capture date: "
            f"{freshness.isoformat() if freshness else 'N/A'}   |   "
            f"Files consolidated: {table.get('upload_count', 0)}"
        )
        ws.cell(row=1, column=1, value=fresh_text)
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(columns))
        banner = ws.cell(row=1, column=1)
        banner.font = Font(bold=True, color="FFFFFF", size=11)
        banner.fill = PatternFill("solid", fgColor="2e7d32")
        banner.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        for col_idx, (_key, ar, en) in enumerate(columns, start=1):
            _styled_header(ws.cell(row=2, column=col_idx), ar, en)
        ws.row_dimensions[2].height = 56
        ws.freeze_panes = "F3"

        indicator_codes = [ind.code for ind in spec.indicators]
        for row_idx, row in enumerate(table.get("rows", []), start=3):
            ws.cell(row=row_idx, column=1,
                    value=_safe_text(row.municipality_id))
            ws.cell(row=row_idx, column=2, value=_safe_text(row.name_ar))
            ws.cell(row=row_idx, column=3, value=_safe_text(row.name_en))
            ws.cell(row=row_idx, column=4,
                    value=(row.capture_date.isoformat()
                           if row.capture_date else ""))
            ws.cell(row=row_idx, column=5, value=_safe_text(row.agency_name))
            col = 6
            for code in indicator_codes:
                ws.cell(row=row_idx, column=col, value=row.metrics.get(code))
                col += 1
            ws.cell(row=row_idx, column=col, value=_safe_text(row.notes))

        width_map = {
            "municipality_id": 14, "name_ar": 26, "name_en": 26,
            "capture_date": 14, "agency_name": 26, "notes": 40,
        }
        for col_idx, (key, *_rest) in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = (
                width_map.get(key, 22)
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Upload reader & comparison table
# --------------------------------------------------------------------------- #

@dataclass
class ConsolidatedRow:
    municipality_id: str
    name_ar: str
    name_en: str
    capture_date: date | None
    agency_name: str | None
    metrics: dict[str, float | None] = field(default_factory=dict)
    notes: str | None = None
    source_file: str = ""


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        d = value.date()
    elif isinstance(value, date):
        d = value
    else:
        try:
            d = datetime.strptime(str(value).strip()[:10], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return None
    # Reject dates outside the plausible window — guards against bypassed
    # Excel validation poisoning the comparison table.
    if d < EARLIEST_CAPTURE_DATE:
        return None
    if d > date.today():
        return None
    return d


def workbook_domain_key(raw: bytes) -> str | None:
    """Read the ``cara_libya_domain_key`` custom property from a workbook
    given its raw bytes. Returns ``None`` if absent or unreadable.

    Used by the upload route to confirm an uploaded file was generated for
    the same domain it is being uploaded to.
    """
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        return None
    try:
        container = wb.custom_doc_props
        items = getattr(container, "props", None) or []
        for prop in items:
            if getattr(prop, "name", None) == DOMAIN_PROPERTY_NAME:
                return str(prop.value) if prop.value is not None else None
    except Exception:
        return None
    return None


def _parse_value(value: Any, ind: Indicator) -> float | None:
    if value is None or value == "":
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    if f < ind.min_value:
        return None
    if ind.max_value is not None and f > ind.max_value:
        return None
    return f


def _read_workbook(path: Path, spec: DomainSpec) -> list[ConsolidatedRow]:
    rows: list[ConsolidatedRow] = []
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        logger.warning("Could not open uploaded workbook %s: %s", path, exc)
        return rows
    ws = wb["Data Entry"] if "Data Entry" in wb.sheetnames else wb.active

    columns = all_columns(spec)
    key_index = {key: i for i, (key, *_rest) in enumerate(columns)}

    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if excel_row is None:
            continue
        padded = list(excel_row) + [None] * (len(columns) - len(excel_row))

        capture = _parse_date(padded[key_index["capture_date"]])
        muni_id = padded[key_index["municipality_id"]]
        if not muni_id or capture is None:
            continue

        metrics: dict[str, float | None] = {}
        for ind in spec.indicators:
            metrics[ind.code] = _parse_value(padded[key_index[ind.code]], ind)

        # A row must contribute at least one valid metric; otherwise
        # admitting it into consolidation lets a later all-null upload
        # silently overwrite previously-valid data on the same capture
        # date (mtime tie-break). Skip defensively.
        if not any(v is not None for v in metrics.values()):
            continue

        notes_val = padded[key_index["notes"]]
        rows.append(ConsolidatedRow(
            municipality_id=str(muni_id).strip(),
            name_ar=str(padded[key_index["name_ar"]] or "").strip(),
            name_en=str(padded[key_index["name_en"]] or "").strip(),
            capture_date=capture,
            agency_name=(str(padded[key_index["agency_name"]]).strip()
                         if padded[key_index["agency_name"]] else None),
            metrics=metrics,
            notes=(str(notes_val).strip() if notes_val else None),
            source_file=path.name,
        ))
    return rows


def consolidated_table(spec: DomainSpec) -> dict[str, Any]:
    """Read every uploaded workbook for this domain and consolidate."""
    upload_dir = upload_dir_for(spec)
    upload_dir.mkdir(parents=True, exist_ok=True)

    files = sorted(upload_dir.glob("*.xlsx"))
    if not files:
        return {
            "rows": [],
            "freshness": None,
            "upload_count": 0,
            "last_upload_at": None,
        }

    # Process files oldest → newest by mtime so that on capture-date ties
    # the row from the most recent upload deterministically wins.
    files = sorted(files, key=lambda p: (os.path.getmtime(p), p.name))
    latest_per_muni: dict[str, ConsolidatedRow] = {}
    for f in files:
        for row in _read_workbook(f, spec):
            existing = latest_per_muni.get(row.municipality_id)
            if existing is None:
                latest_per_muni[row.municipality_id] = row
                continue
            if (row.capture_date and existing.capture_date
                    and row.capture_date >= existing.capture_date):
                latest_per_muni[row.municipality_id] = row

    rows = sorted(latest_per_muni.values(),
                  key=lambda r: r.name_en or r.municipality_id)
    freshness = max((r.capture_date for r in rows if r.capture_date),
                    default=None)
    last_upload_at = datetime.fromtimestamp(
        max(os.path.getmtime(f) for f in files)
    )
    return {
        "rows": rows,
        "freshness": freshness,
        "upload_count": len(files),
        "last_upload_at": last_upload_at,
    }


# --------------------------------------------------------------------------- #
# Export consolidated table
# --------------------------------------------------------------------------- #

def build_export_workbook(spec: DomainSpec, table: dict[str, Any]) -> bytes:
    """Export the consolidated comparison table back to .xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    ws.sheet_view.rightToLeft = True

    columns = all_columns(spec)

    # Freshness banner (merged across all columns).
    freshness = table.get("freshness")
    fresh_text = (
        f"{spec.name_ar}   |   {spec.name_en}   ||   "
        f"تاريخ أحدث البيانات: {freshness.isoformat() if freshness else 'غير متاح'}   |   "
        f"Latest capture date: {freshness.isoformat() if freshness else 'N/A'}   |   "
        f"Files consolidated: {table.get('upload_count', 0)}"
    )
    ws.cell(row=1, column=1, value=fresh_text)
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(columns))
    banner = ws.cell(row=1, column=1)
    banner.font = Font(bold=True, color="FFFFFF", size=11)
    banner.fill = PatternFill("solid", fgColor="2e7d32")
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Header row at row 2.
    for col_idx, (_key, ar, en) in enumerate(columns, start=1):
        _styled_header(ws.cell(row=2, column=col_idx), ar, en)
    ws.row_dimensions[2].height = 56
    ws.freeze_panes = "F3"

    # Data rows. All free-text fields go through _safe_text().
    indicator_codes = [ind.code for ind in spec.indicators]
    for row_idx, row in enumerate(table.get("rows", []), start=3):
        ws.cell(row=row_idx, column=1, value=_safe_text(row.municipality_id))
        ws.cell(row=row_idx, column=2, value=_safe_text(row.name_ar))
        ws.cell(row=row_idx, column=3, value=_safe_text(row.name_en))
        ws.cell(row=row_idx, column=4,
                value=row.capture_date.isoformat() if row.capture_date else "")
        ws.cell(row=row_idx, column=5, value=_safe_text(row.agency_name))
        col = 6
        for code in indicator_codes:
            ws.cell(row=row_idx, column=col, value=row.metrics.get(code))
            col += 1
        ws.cell(row=row_idx, column=col, value=_safe_text(row.notes))

    # Column widths.
    width_map = {
        "municipality_id": 14, "name_ar": 26, "name_en": 26,
        "capture_date": 14, "agency_name": 26, "notes": 40,
    }
    for col_idx, (key, *_rest) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            width_map.get(key, 22)
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
